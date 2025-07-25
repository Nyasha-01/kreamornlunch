from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
import json
import os
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import io
import hashlib
import threading
import time

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'

# File to store menu data
MENU_FILE = 'weekly_menu.json'
ORDERS_FILE = 'orders.json'
CATEGORIES_FILE = 'categories.json'
USERS_FILE = 'users.json'
EXCLUSIVE_CATEGORIES_FILE = 'exclusive_categories.json'
MEAL_PRICE_FILE = 'meal_price.json'

# Days of the week (weekdays only)
DAYS_OF_WEEK = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

# Default categories
DEFAULT_CATEGORIES = ['proteins', 'starch', 'salad']

def get_current_day():
    """Get the current day of the week"""
    return datetime.now().strftime('%A')

def load_categories():
    """Load categories from file"""
    if os.path.exists(CATEGORIES_FILE):
        with open(CATEGORIES_FILE, 'r') as f:
            return json.load(f)
    return DEFAULT_CATEGORIES.copy()

def save_categories(categories):
    """Save categories to file"""
    with open(CATEGORIES_FILE, 'w') as f:
        json.dump(categories, f, indent=2)

def load_exclusive_categories():
    """Load exclusive categories from file"""
    if os.path.exists(EXCLUSIVE_CATEGORIES_FILE):
        with open(EXCLUSIVE_CATEGORIES_FILE, 'r') as f:
            return json.load(f)
    return []

def save_exclusive_categories(exclusive_categories):
    """Save exclusive categories to file"""
    with open(EXCLUSIVE_CATEGORIES_FILE, 'w') as f:
        json.dump(exclusive_categories, f, indent=2)

def load_meal_price():
    """Load meal price from file"""
    if os.path.exists(MEAL_PRICE_FILE):
        with open(MEAL_PRICE_FILE, 'r') as f:
            data = json.load(f)
            return data.get('price', 0.0)
    return 0.0

def save_meal_price(price):
    """Save meal price to file"""
    with open(MEAL_PRICE_FILE, 'w') as f:
        json.dump({'price': float(price)}, f, indent=2)

def load_weekly_menu():
    """Load the weekly menu from file"""
    categories = load_categories()
    if os.path.exists(MENU_FILE):
        with open(MENU_FILE, 'r') as f:
            menu = json.load(f)
            # Ensure all categories exist for all days
            for day in DAYS_OF_WEEK:
                if day not in menu:
                    menu[day] = {}
                for category in categories:
                    if category not in menu[day]:
                        menu[day][category] = []
            return menu
    # Initialize with empty menu for each day
    return {day: {category: [] for category in categories} for day in DAYS_OF_WEEK}

def save_weekly_menu(menu):
    """Save the weekly menu to file"""
    with open(MENU_FILE, 'w') as f:
        json.dump(menu, f, indent=2)

def get_today_menu():
    """Get today's menu"""
    weekly_menu = load_weekly_menu()
    current_day = get_current_day()
    categories = load_categories()
    return weekly_menu.get(current_day, {category: [] for category in categories})

def load_orders():
    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_orders(orders):
    with open(ORDERS_FILE, 'w') as f:
        json.dump(orders, f, indent=2)

def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            return json.load(f)

    # Create default admin user if no users exist
    default_users = {
        "admin": {
            'first_name': 'Admin',
            'last_name': 'User',
            'full_name': 'admin',
            'password': hash_password('admin1234'),
            'is_admin': True,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    }
    save_users(default_users)
    return default_users

def save_users(users):
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f, indent=2)

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, hashed):
    return hash_password(password) == hashed

def cleanup_old_orders():
    """Remove orders older than 15 days to prevent storage buildup"""
    try:
        orders = load_orders()
        cutoff_date = datetime.now() - timedelta(days=15)

        dates_to_remove = []
        for date_str in orders.keys():
            try:
                order_date = datetime.strptime(date_str, '%Y-%m-%d')
                if order_date < cutoff_date:
                    dates_to_remove.append(date_str)
            except ValueError:
                # Skip invalid date formats
                continue

        # Remove old orders
        for date_str in dates_to_remove:
            del orders[date_str]

        if dates_to_remove:
            save_orders(orders)
            print(f"Cleaned up {len(dates_to_remove)} old order dates: {dates_to_remove}")

    except Exception as e:
        print(f"Error during cleanup: {e}")

def daily_cleanup_task():
    """Background task that runs cleanup daily"""
    while True:
        try:
            # Run cleanup every 24 hours
            time.sleep(24 * 60 * 60)  # 24 hours in seconds
            cleanup_old_orders()
        except Exception as e:
            print(f"Error in daily cleanup task: {e}")
            time.sleep(60)  # Wait 1 minute before retrying

def start_cleanup_scheduler():
    """Start the background cleanup scheduler"""
    cleanup_thread = threading.Thread(target=daily_cleanup_task, daemon=True)
    cleanup_thread.start()
    print("Started automatic 15-day cleanup scheduler")

def is_logged_in():
    return 'user_id' in session

def get_current_user():
    if is_logged_in():
        users = load_users()
        return users.get(session['user_id'])
    return None

def is_admin():
    user = get_current_user()
    return user and user.get('is_admin', False)

def admin_required(f):
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Please log in to access this page')
            return redirect(url_for('login'))
        if not is_admin():
            flash('Admin access required')
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
def home():
    return render_template('index.html', user=get_current_user())

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not all([first_name, last_name, password]):
            flash('All fields are required')
            return render_template('register.html')

        if password != confirm_password:
            flash('Passwords do not match')
            return render_template('register.html')

        if len(password) < 6:
            flash('Password must be at least 6 characters long')
            return render_template('register.html')

        users = load_users()
        full_name = f"{first_name} {last_name}"

        # Check if user already exists
        for user_id, user_data in users.items():
            if user_data['full_name'].lower() == full_name.lower():
                flash('A user with this name already exists')
                return render_template('register.html')

        # Create new user
        user_id = str(len(users) + 1)
        # First user is automatically admin
        is_first_user = len(users) == 0
        users[user_id] = {
            'first_name': first_name,
            'last_name': last_name,
            'full_name': full_name,
            'password': hash_password(password),
            'is_admin': is_first_user,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        save_users(users)
        session['user_id'] = user_id
        flash('Registration successful! You are now logged in.')
        # Check if user is admin and redirect accordingly
        if is_first_user:
            return redirect(url_for('admin_page'))
        else:
            return redirect(url_for('order_page'))

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        password = request.form.get('password', '')

        if not all([full_name, password]):
            flash('All fields are required')
            return render_template('login.html')

        users = load_users()

        # Find user by full name or username
        user_id = None
        for uid, user_data in users.items():
            if (user_data['full_name'].lower() == full_name.lower() or 
                uid.lower() == full_name.lower()):
                if verify_password(password, user_data['password']):
                    user_id = uid
                    break

        if user_id:
            session['user_id'] = user_id
            flash('Login successful!')
            # Check if user is admin and redirect accordingly
            user = users[user_id]
            if user.get('is_admin', False):
                return redirect(url_for('admin_page'))
            else:
                return redirect(url_for('order_page'))
        else:
            flash('Invalid name or password')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out')
    return redirect(url_for('home'))

@app.route('/order')
def order_page():
    # Redirect to weekly order page instead
    return redirect(url_for('weekly_order_page'))

@app.route('/weekly_order')
def weekly_order_page():
    if not is_logged_in():
        flash('Please log in to place orders')
        return redirect(url_for('login'))

    weekly_menu = load_weekly_menu()
    categories = load_categories()
    exclusive_categories = load_exclusive_categories()
    user = get_current_user()
    
    # Debug logging for menu delivery
    print(f"=== WEEKLY ORDER PAGE DEBUG ===")
    print(f"User: {user['full_name'] if user else 'None'}")
    print(f"Categories: {categories}")
    print(f"Exclusive categories: {exclusive_categories}")
    print(f"Total menu items: {sum(len(items) for day_menu in weekly_menu.values() for items in day_menu.values())}")
    
    # Check each day's menu content
    for day in DAYS_OF_WEEK:
        day_menu = weekly_menu.get(day, {})
        day_total = sum(len(items) for items in day_menu.values())
        print(f"  {day}: {day_total} total items")
        for category in categories:
            items = day_menu.get(category, [])
            print(f"    {category}: {len(items)} items - {items[:3]}{'...' if len(items) > 3 else ''}")
    
    print(f"=== END WEEKLY ORDER PAGE DEBUG ===\n")
    
    return render_template('weekly_order.html', weekly_menu=weekly_menu, days=DAYS_OF_WEEK, categories=categories, exclusive_categories=exclusive_categories, user=user, current_day=get_current_day())

# Removed single-day order submission - using weekly orders only

@app.route('/submit_daily_order', methods=['POST'])
def submit_daily_order():
    if not is_logged_in():
        flash('Please log in to place an order')
        return redirect(url_for('login'))

    user = get_current_user()
    user_name = user['full_name']
    categories = load_categories()
    order_day = request.form.get('order_day')

    if not order_day or order_day not in DAYS_OF_WEEK:
        flash('Invalid day selected')
        return redirect(url_for('weekly_order_page'))

    # Calculate the date for the selected day
    today = datetime.now()
    days_since_monday = today.weekday()
    monday = today - timedelta(days=days_since_monday)
    day_index = DAYS_OF_WEEK.index(order_day)
    order_date = (monday + timedelta(days=day_index)).strftime('%Y-%m-%d')

    orders = load_orders()
    if order_date not in orders:
        orders[order_date] = {}

    # Build order dict dynamically based on categories
    order_data = {
        'date_ordered': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'user_id': session['user_id'],
        'single_day_order': True
    }

    # Get selections for the specific day
    has_selections = False
    for category in categories:
        selection = request.form.get(f'{category}_{order_day}')
        order_data[category] = selection
        if selection:
            has_selections = True

    if has_selections:
        orders[order_date][user_name] = order_data
        save_orders(orders)
        flash(f'Your order for {order_day} has been submitted successfully!')
    else:
        flash(f'No items selected for {order_day}. Please select at least one item.')

    return redirect(url_for('weekly_order_page'))

@app.route('/submit_weekly_order', methods=['POST'])
def submit_weekly_order():
    if not is_logged_in():
        flash('Please log in to place orders')
        return redirect(url_for('login'))

    user = get_current_user()
    user_name = user['full_name']
    categories = load_categories()
    orders = load_orders()

    # Get the starting date (Monday of current week)
    today = datetime.now()
    days_since_monday = today.weekday()
    monday = today - timedelta(days=days_since_monday)

    orders_placed = 0

    for i, day in enumerate(DAYS_OF_WEEK):
        # Calculate the date for this day
        order_date = (monday + timedelta(days=i)).strftime('%Y-%m-%d')

        # Check if any items were selected for this day
        has_selections = False
        order_data = {
            'date_ordered': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'user_id': session['user_id'],
            'weekly_order': True
        }

        for category in categories:
            selection = request.form.get(f'{category}_{day}')
            order_data[category] = selection
            if selection:
                has_selections = True

        # Only save order if user made selections for this day
        if has_selections:
            if order_date not in orders:
                orders[order_date] = {}

            orders[order_date][user_name] = order_data
            orders_placed += 1

    save_orders(orders)

    if orders_placed > 0:
        flash(f'Your weekly orders have been submitted successfully! {orders_placed} day(s) ordered.')
    else:
        flash('No orders were placed. Please select items for at least one day.')

    return redirect(url_for('weekly_order_page'))

@app.route('/admin')
@admin_required
def admin_page():
    weekly_menu = load_weekly_menu()
    orders = load_orders()

    # Get selected date from query parameter, default to today
    selected_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    selected_orders = orders.get(selected_date, {})
    current_day = get_current_day()

    # Generate list of dates for the dropdown (last 7 days and next 7 days)
    today = datetime.now()
    date_options = []
    for i in range(-7, 8):
        date = today + timedelta(days=i)
        date_str = date.strftime('%Y-%m-%d')
        day_name = date.strftime('%A')
        date_options.append({
            'date': date_str,
            'display': f"{day_name}, {date.strftime('%B %d, %Y')}",
            'is_today': i == 0
        })

    categories = load_categories()
    all_users = load_users()
    exclusive_categories = load_exclusive_categories()
    meal_price = load_meal_price()

    return render_template('admin.html', 
                         weekly_menu=weekly_menu, 
                         orders=selected_orders, 
                         selected_date=selected_date,
                         date_options=date_options,
                         days=DAYS_OF_WEEK,
                         current_day=current_day,
                         categories=categories,
                         all_users=all_users,
                         exclusive_categories=exclusive_categories,
                         meal_price=meal_price)

@app.route('/update_weekly_menu', methods=['POST'])
@admin_required
def update_weekly_menu():
    print(f"=== MENU UPDATE DEBUG START ===")
    print(f"Request method: {request.method}")
    print(f"Form data keys: {list(request.form.keys())}")
    print(f"Form data: {dict(request.form)}")
    
    weekly_menu = load_weekly_menu()
    categories = load_categories()
    current_section = request.form.get('current_section', 'menu-management')
    
    # Get all unique categories across all days
    all_categories = set()
    for day_menu in weekly_menu.values():
        all_categories.update(day_menu.keys())
    all_categories = sorted(list(all_categories))
    
    print(f"Global categories: {categories}")
    print(f"All categories found in menu: {all_categories}")
    print(f"Current menu before update: {weekly_menu}")

    # Track if any changes were made
    changes_made = False
    change_details = []

    for day in DAYS_OF_WEEK:
        print(f"\nProcessing day: {day}")
        # Process all categories that exist for this specific day
        day_categories = list(weekly_menu.get(day, {}).keys())
        print(f"Categories for {day}: {day_categories}")
        
        for category in day_categories:
            form_key = f'{category}_{day}'
            items = request.form.getlist(form_key)
            print(f"  Form key '{form_key}': {items}")
            
            # Filter out empty entries
            items = [item.strip() for item in items if item.strip()]
            print(f"  Filtered items: {items}")

            # Check if this is different from current menu
            current_items = weekly_menu.get(day, {}).get(category, [])
            print(f"  Current items in menu: {current_items}")
            
            if items != current_items:
                changes_made = True
                change_detail = f"{day} - {category}: {current_items} -> {items}"
                change_details.append(change_detail)
                print(f"  CHANGE DETECTED: {change_detail}")

            weekly_menu[day][category] = items

    print(f"\nTotal changes made: {len(change_details)}")
    for detail in change_details:
        print(f"  - {detail}")

    print(f"\nSaving menu to file...")
    save_weekly_menu(weekly_menu)

    # Verify the file was written correctly
    try:
        verify_menu = load_weekly_menu()
        print(f"Menu verification successful. Items in file: {sum(len(items) for day_menu in verify_menu.values() for items in day_menu.values())}")
        
        # Compare with what we just saved
        if verify_menu == weekly_menu:
            print("✓ Verified menu matches what we saved")
        else:
            print("✗ WARNING: Verified menu does not match what we saved!")
            print(f"Saved: {weekly_menu}")
            print(f"Loaded: {verify_menu}")
            
    except Exception as e:
        print(f"Error verifying menu save: {e}")

    # Check file permissions and size
    try:
        file_stat = os.stat(MENU_FILE)
        print(f"Menu file size: {file_stat.st_size} bytes")
        print(f"Menu file last modified: {datetime.fromtimestamp(file_stat.st_mtime)}")
    except Exception as e:
        print(f"Error checking file stats: {e}")

    # Trigger notification for menu update only if changes were made
    if changes_made:
        try:
            from notification_service import get_notification_service
            notification_service = get_notification_service(request.host_url)
            if notification_service:
                notification_service.send_menu_upload_notification()
                print("Menu update notification sent successfully")
        except Exception as e:
            print(f"Error sending menu update notification: {e}")

        flash(f'Weekly menu updated successfully! {len(change_details)} changes made. All users will now see the new menu.')
    else:
        flash('No changes detected in the menu.')

    print(f"=== MENU UPDATE DEBUG END ===\n")
    return redirect(url_for('admin_page', section=current_section))

@app.route('/admin_place_order', methods=['POST'])
@admin_required
def admin_place_order():
    user_name = request.form.get('user_name')
    order_date = request.form.get('order_date', datetime.now().strftime('%Y-%m-%d'))
    current_section = request.form.get('current_section', 'place-order')
    categories = load_categories()

    if not user_name:
        flash('Please enter the person\'s name')
        return redirect(url_for('admin_page', section=current_section))

    orders = load_orders()

    if order_date not in orders:
        orders[order_date] = {}

    # Build order dict dynamically based on categories
    order_data = {
        'date_ordered': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'placed_by_admin': True
    }

    for category in categories:
        order_data[category] = request.form.get(category)

    orders[order_date][user_name] = order_data

    save_orders(orders)
    flash(f'Order placed successfully for {user_name}!')
    return redirect(url_for('admin_page', date=order_date, section=current_section))

@app.route('/add_category', methods=['POST'])
@admin_required
def add_category():
    category_name = request.form.get('category_name', '').strip().lower()
    category_scope = request.form.get('category_scope', 'all_days')
    current_section = request.form.get('current_section', 'categories')

    print(f"=== ADD CATEGORY DEBUG START ===")
    print(f"Raw form data: {dict(request.form)}")
    print(f"Category name: '{category_name}'")
    print(f"Category scope: '{category_scope}'")
    print(f"Current section: '{current_section}'")

    if not category_name:
        print("ERROR: Empty category name")
        flash('Please enter a category name')
        return redirect(url_for('admin_page', section=current_section))

    categories = load_categories()
    weekly_menu = load_weekly_menu()
    
    print(f"Before - Global categories: {categories}")
    print(f"Before - Weekly menu structure: {[(day, list(day_menu.keys())) for day, day_menu in weekly_menu.items()]}")
    
    # Ensure weekly_menu has all days initialized
    for day in DAYS_OF_WEEK:
        if day not in weekly_menu:
            weekly_menu[day] = {}
            print(f"Initialized empty day: {day}")

    if category_scope == 'all_days':
        print(f"Processing 'all_days' scope")
        # Check if category already exists in global categories
        if category_name in categories:
            print(f"Category '{category_name}' already exists in global categories")
            flash(f'Category "{category_name}" already exists in global categories')
            return redirect(url_for('admin_page', section=current_section))
        
        # Add to global categories
        categories.append(category_name)
        save_categories(categories)
        print(f"Added '{category_name}' to global categories")
        
        # Add to all days
        for day in DAYS_OF_WEEK:
            if category_name not in weekly_menu[day]:
                weekly_menu[day][category_name] = []
                print(f"Added '{category_name}' to day {day}")
            else:
                print(f"Category '{category_name}' already exists in {day}")
        
        flash(f'Category "{category_name}" added to all days successfully!')
    else:
        # Add to specific day only
        print(f"Processing specific day scope: {category_scope}")
        if category_scope not in DAYS_OF_WEEK:
            print(f"ERROR: Invalid day selection: '{category_scope}' not in {DAYS_OF_WEEK}")
            flash('Invalid day selection')
            return redirect(url_for('admin_page', section=current_section))
        
        # Check if category already exists in that specific day
        if category_name in weekly_menu[category_scope]:
            print(f"Category '{category_name}' already exists in {category_scope}")
            flash(f'Category "{category_name}" already exists in {category_scope}')
            return redirect(url_for('admin_page', section=current_section))
        
        # Add to the specific day only
        weekly_menu[category_scope][category_name] = []
        print(f"Added '{category_name}' to {category_scope}")
        print(f"Day {category_scope} now has categories: {list(weekly_menu[category_scope].keys())}")
        
        flash(f'Category "{category_name}" added to {category_scope} only!')
    
    # Save the updated weekly menu
    print(f"Saving weekly menu...")
    save_weekly_menu(weekly_menu)
    
    # Verify save worked
    print(f"Verifying save...")
    verify_menu = load_weekly_menu()
    print(f"After save - Global categories: {load_categories()}")
    print(f"After save - Weekly menu structure: {[(day, list(day_menu.keys())) for day, day_menu in verify_menu.items()]}")
    
    if category_scope != 'all_days' and category_scope in DAYS_OF_WEEK:
        verify_categories = list(verify_menu.get(category_scope, {}).keys())
        print(f"Verification - {category_scope} categories: {verify_categories}")
        if category_name in verify_categories:
            print(f"✓ SUCCESS: Category '{category_name}' successfully added to {category_scope}")
        else:
            print(f"✗ FAILED: Category '{category_name}' not found in {category_scope} after save")
    
    print(f"=== ADD CATEGORY DEBUG END ===\n")
    return redirect(url_for('admin_page', section=current_section))

@app.route('/remove_category', methods=['POST'])
@admin_required
def remove_category():
    # Redirect to the new specific routes
    return redirect(url_for('admin_page', section='categories'))

@app.route('/remove_global_category', methods=['POST'])
@admin_required
def remove_global_category():
    category_name = request.form.get('category_name')
    current_section = request.form.get('current_section', 'categories')

    if not category_name:
        flash('Invalid category')
        return redirect(url_for('admin_page', section=current_section))

    categories = load_categories()

    if category_name in DEFAULT_CATEGORIES:
        flash(f'Cannot remove default category "{category_name}"')
        return redirect(url_for('admin_page', section=current_section))

    if category_name not in categories:
        flash(f'Category "{category_name}" not found in global categories')
        return redirect(url_for('admin_page', section=current_section))

    # Remove from global categories
    categories.remove(category_name)
    save_categories(categories)

    # Remove category from weekly menu (all days)
    weekly_menu = load_weekly_menu()
    for day in DAYS_OF_WEEK:
        if category_name in weekly_menu[day]:
            del weekly_menu[day][category_name]
    save_weekly_menu(weekly_menu)

    # Remove from exclusive categories if exists
    exclusive_categories = load_exclusive_categories()
    if category_name in exclusive_categories:
        exclusive_categories.remove(category_name)
        save_exclusive_categories(exclusive_categories)

    flash(f'Global category "{category_name}" removed from all days successfully!')
    return redirect(url_for('admin_page', section=current_section))

@app.route('/remove_day_category', methods=['POST'])
@admin_required
def remove_day_category():
    category_name = request.form.get('category_name')
    day_name = request.form.get('day_name')
    current_section = request.form.get('current_section', 'categories')

    if not category_name or not day_name:
        flash('Invalid category or day')
        return redirect(url_for('admin_page', section=current_section))

    if day_name not in DAYS_OF_WEEK:
        flash('Invalid day')
        return redirect(url_for('admin_page', section=current_section))

    weekly_menu = load_weekly_menu()
    
    if day_name not in weekly_menu or category_name not in weekly_menu[day_name]:
        flash(f'Category "{category_name}" not found in {day_name}')
        return redirect(url_for('admin_page', section=current_section))

    # Remove category from specific day only
    del weekly_menu[day_name][category_name]
    save_weekly_menu(weekly_menu)

    flash(f'Category "{category_name}" removed from {day_name} successfully!')
    return redirect(url_for('admin_page', section=current_section))

@app.route('/toggle_exclusive_category', methods=['POST'])
@admin_required
def toggle_exclusive_category():
    category_name = request.form.get('category_name')
    action = request.form.get('action')

    if not category_name or action not in ['make_exclusive', 'make_normal']:
        flash('Invalid request')
        return redirect(url_for('admin_page'))

    categories = load_categories()
    if category_name not in categories:
        flash('Category not found')
        return redirect(url_for('admin_page'))

    exclusive_categories = load_exclusive_categories()

    if action == 'make_exclusive':
        if category_name not in exclusive_categories:
            exclusive_categories.append(category_name)
            flash(f'Category "{category_name}" is now exclusive - selecting from it will disable other categories')
    else:
        if category_name in exclusive_categories:
            exclusive_categories.remove(category_name)
            flash(f'Category "{category_name}" is no longer exclusive')

    save_exclusive_categories(exclusive_categories)
    return redirect(url_for('admin_page'))

@app.route('/toggle_admin', methods=['POST'])
@admin_required
def toggle_admin():
    user_id = request.form.get('user_id')
    action = request.form.get('action')

    if not user_id or action not in ['promote', 'demote']:
        flash('Invalid request')
        return redirect(url_for('admin_page'))

    users = load_users()

    if user_id not in users:
        flash('User not found')
        return redirect(url_for('admin_page'))

    # Prevent demoting yourself if you're the only admin
    current_user = get_current_user()
    if action == 'demote' and user_id == session['user_id']:
        admin_count = sum(1 for u in users.values() if u.get('is_admin', False))
        if admin_count <= 1:
            flash('Cannot demote yourself - you are the only admin')
            return redirect(url_for('admin_page'))

    if action == 'promote':
        users[user_id]['is_admin'] = True
        flash(f'{users[user_id]["full_name"]} has been promoted to admin')
    else:
        users[user_id]['is_admin'] = False
        flash(f'{users[user_id]["full_name"]} has been demoted from admin')

    save_users(users)
    return redirect(url_for('admin_page'))

@app.route('/reset_password', methods=['POST'])
@admin_required
def reset_password():
    user_id = request.form.get('user_id')
    new_password = request.form.get('new_password')

    if not user_id or not new_password:
        flash('User ID and new password are required')
        return redirect(url_for('admin_page'))

    if len(new_password) < 6:
        flash('Password must be at least 6 characters long')
        return redirect(url_for('admin_page'))

    users = load_users()

    if user_id not in users:
        flash('User not found')
        return redirect(url_for('admin_page'))

    # Update the user's password
    users[user_id]['password'] = hash_password(new_password)
    save_users(users)

    flash(f'Password has been reset for {users[user_id]["full_name"]}')
    return redirect(url_for('admin_page'))

@app.route('/update_meal_price', methods=['POST'])
@admin_required
def update_meal_price():
    meal_price = request.form.get('meal_price')
    current_section = request.form.get('current_section', 'orders-management')
    
    if not meal_price:
        flash('Please enter a meal price')
        return redirect(url_for('admin_page', section=current_section))
    
    try:
        price_float = float(meal_price)
        if price_float < 0:
            flash('Meal price cannot be negative')
            return redirect(url_for('admin_page', section=current_section))
        
        save_meal_price(price_float)
        flash(f'Meal price updated to ${price_float:.2f}')
    except ValueError:
        flash('Invalid meal price format')
    
    return redirect(url_for('admin_page', section=current_section))

@app.route('/admin_create_account', methods=['POST'])
@admin_required
def admin_create_account():
    first_name = request.form.get('first_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    password = request.form.get('password', '')
    is_admin = request.form.get('is_admin') == 'on'

    if not all([first_name, last_name, password]):
        flash('All fields are required')
        return redirect(url_for('admin_page'))

    if len(password) < 6:
        flash('Password must be at least 6 characters long')
        return redirect(url_for('admin_page'))

    users = load_users()
    full_name = f"{first_name} {last_name}"

    # Check if user already exists
    for user_id, user_data in users.items():
        if user_data['full_name'].lower() == full_name.lower():
            flash('A user with this name already exists')
            return redirect(url_for('admin_page'))

    # Create new user
    user_id = str(max([int(uid) for uid in users.keys() if uid.isdigit()] + [0]) + 1)
    users[user_id] = {
        'first_name': first_name,
        'last_name': last_name,
        'full_name': full_name,
        'password': hash_password(password),
        'is_admin': is_admin,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    save_users(users)
    flash(f'Account created successfully for {full_name}')
    return redirect(url_for('admin_page'))

@app.route('/delete_account', methods=['POST'])
@admin_required
def delete_account():
    user_id = request.form.get('user_id')

    if not user_id:
        flash('Invalid user ID')
        return redirect(url_for('admin_page'))

    users = load_users()

    if user_id not in users:
        flash('User not found')
        return redirect(url_for('admin_page'))

    # Prevent deleting yourself if you're the only admin
    current_user = get_current_user()
    if user_id == session['user_id']:
        admin_count = sum(1 for u in users.values() if u.get('is_admin', False))
        if admin_count <= 1:
            flash('Cannot delete yourself - you are the only admin')
            return redirect(url_for('admin_page'))

    user_name = users[user_id]['full_name']
    del users[user_id]
    save_users(users)

    # Remove user's orders from all dates
    orders = load_orders()
    for date in orders:
        if user_name in orders[date]:
            del orders[date][user_name]
    save_orders(orders)

    flash(f'Account deleted for {user_name}')
    return redirect(url_for('admin_page'))

@app.route('/delete_order', methods=['POST'])
@admin_required
def delete_order():
    order_date = request.form.get('order_date')
    user_name = request.form.get('user_name')

    if not order_date or not user_name:
        flash('Invalid order information')
        return redirect(url_for('admin_page'))

    orders = load_orders()

    if order_date in orders and user_name in orders[order_date]:
        del orders[order_date][user_name]
        save_orders(orders)
        flash(f'Order deleted for {user_name} on {order_date}')
    else:
        flash('Order not found')

    return redirect(url_for('admin_page', date=order_date))

# Removed weekly menu API endpoint

@app.route('/debug/menu')
@admin_required
def debug_menu():
    """Debug endpoint to check current menu state"""
    try:
        weekly_menu = load_weekly_menu()
        categories = load_categories()

        # Read the actual file content
        with open(MENU_FILE, 'r') as f:
            file_content = json.load(f)

        # Check file timestamps
        menu_file_stat = os.stat(MENU_FILE) if os.path.exists(MENU_FILE) else None
        
        debug_info = {
            'loaded_menu': weekly_menu,
            'file_content': file_content,
            'categories': categories,
            'menu_file_exists': os.path.exists(MENU_FILE),
            'menu_file_size': menu_file_stat.st_size if menu_file_stat else 0,
            'menu_last_modified': datetime.fromtimestamp(menu_file_stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S') if menu_file_stat else None,
            'current_day': get_current_day(),
            'total_menu_items': sum(len(items) for day_menu in weekly_menu.values() for items in day_menu.values()),
            'menu_by_day_count': {day: sum(len(items) for items in day_menu.values()) for day, day_menu in weekly_menu.items()}
        }

        return jsonify(debug_info)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/debug/menu_delivery')
def debug_menu_delivery():
    """Debug endpoint to check what menu data is being sent to users"""
    try:
        weekly_menu = load_weekly_menu()
        categories = load_categories()
        exclusive_categories = load_exclusive_categories()
        
        debug_info = {
            'is_logged_in': is_logged_in(),
            'current_user': get_current_user()['full_name'] if is_logged_in() else None,
            'weekly_menu': weekly_menu,
            'categories': categories,
            'exclusive_categories': exclusive_categories,
            'days': DAYS_OF_WEEK,
            'menu_empty_days': [day for day, menu in weekly_menu.items() if all(len(items) == 0 for items in menu.values())],
            'template_data_structure': {
                'weekly_menu_keys': list(weekly_menu.keys()),
                'first_day_categories': list(weekly_menu.get(DAYS_OF_WEEK[0], {}).keys()) if weekly_menu else [],
                'sample_menu_items': {day: {cat: items[:2] for cat, items in day_menu.items()} for day, day_menu in list(weekly_menu.items())[:2]}
            }
        }
        
        return jsonify(debug_info)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export_orders')
@admin_required
def export_orders():
    selected_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    orders = load_orders()
    selected_orders = orders.get(selected_date, {})

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Orders_{selected_date}"

    # Add headers
    categories = load_categories()
    meal_price = load_meal_price()
    headers = ['Name'] + [cat.title() for cat in categories] + ['Meal Price', 'Date Ordered']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Add data
    row = 2
    for name, order in selected_orders.items():
        ws.cell(row=row, column=1, value=name)

        # Add category data
        col = 2
        for category in categories:
            ws.cell(row=row, column=col, value=order.get(category, 'Not selected'))
            col += 1

        # Add meal price
        ws.cell(row=row, column=col, value=f"${meal_price:.2f}" if meal_price > 0 else "Not Set")
        col += 1
        
        # Add date ordered
        ws.cell(row=row, column=col, value=order.get('date_ordered', order.get('timestamp', 'N/A')))
        row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"lunch_orders_{selected_date}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/cleanup', methods=['POST'])
@admin_required
def manual_cleanup():
    """Manually trigger cleanup of old orders"""
    cleanup_old_orders()
    flash('Manual cleanup completed - removed orders older than 15 days')
    return redirect(url_for('admin_page'))

@app.route('/admin/storage_info')
@admin_required
def storage_info():
    """Display storage information"""
    orders = load_orders()
    users = load_users()

    # Count total orders and dates
    total_dates = len(orders)
    total_orders = sum(len(orders_for_date) for orders_for_date in orders.values())
    total_users = len(users)

    # Calculate oldest and newest order dates
    date_strings = list(orders.keys())
    if date_strings:
        try:
            dates = [datetime.strptime(d, '%Y-%m-%d') for d in date_strings]
            oldest_date = min(dates).strftime('%Y-%m-%d')
            newest_date = max(dates).strftime('%Y-%m-%d')
        except ValueError:
            oldest_date = "Invalid date format"
            newest_date = "Invalid date format"
    else:
        oldest_date = "No orders"
        newest_date = "No orders"

    info = {
        'total_users': total_users,
        'total_order_dates': total_dates,
        'total_orders': total_orders,
        'oldest_order_date': oldest_date,
        'newest_order_date': newest_date
    }

    flash(f'Storage Info: {total_users} users, {total_dates} order dates, {total_orders} total orders. Range: {oldest_date} to {newest_date}')
    return redirect(url_for('admin_page'))

@app.route('/upload_menu_pdf', methods=['POST'])
@admin_required
def upload_menu_pdf():
    current_section = request.form.get('current_section', 'menu-management')
    
    if 'pdf_file' not in request.files:
        flash('No PDF file selected')
        return redirect(url_for('admin_page', section=current_section))
    
    file = request.files['pdf_file']
    if file.filename == '':
        flash('No PDF file selected')
        return redirect(url_for('admin_page', section=current_section))
    
    if not file.filename.lower().endswith('.pdf'):
        flash('Please upload a PDF file')
        return redirect(url_for('admin_page', section=current_section))
    
    target_day = request.form.get('target_day', 'ALL_DAYS')
    
    try:
        try:
            import PyPDF2
        except ImportError:
            flash('PyPDF2 library not installed. Please install it to use PDF upload feature.')
            return redirect(url_for('admin_page', section=current_section))
        
        # Read PDF content
        pdf_reader = PyPDF2.PdfReader(file)
        text_content = ""
        
        for page in pdf_reader.pages:
            text_content += page.extract_text() + "\n"
        
        if not text_content.strip():
            flash('Could not extract text from PDF. Please ensure the PDF contains readable text.')
            return redirect(url_for('admin_page', section=current_section))
        
        print(f"=== PDF PROCESSING DEBUG ===")
        print(f"Raw text length: {len(text_content)}")
        print(f"First 500 chars: {text_content[:500]}")
        
        # Process extracted text
        weekly_menu_items = process_pdf_text(text_content)
        
        print(f"Processed weekly menu items: {weekly_menu_items}")
        print(f"=== END PDF PROCESSING DEBUG ===")
        
        # Check if any items were extracted
        total_items = sum(len(items) for day_menu in weekly_menu_items.values() for items in day_menu.values())
        if total_items == 0:
            flash('No menu items could be extracted from the PDF. Please check the PDF format.')
            return redirect(url_for('admin_page', section=current_section))
        
        # Load existing menu and categories (only work with existing categories)
        weekly_menu = load_weekly_menu()
        categories = load_categories()
        
        # Only process the three main categories: proteins, starch, salad
        main_categories = ['proteins', 'starch', 'salad']
        
        # Filter weekly_menu_items to only include main categories
        filtered_menu_items = {}
        for day in DAYS_OF_WEEK:
            filtered_menu_items[day] = {}
            for category in main_categories:
                if category in weekly_menu_items.get(day, {}):
                    filtered_menu_items[day][category] = weekly_menu_items[day][category]
                else:
                    filtered_menu_items[day][category] = []
        
        # Replace weekly_menu_items with filtered version
        weekly_menu_items = filtered_menu_items
        
        items_added = 0
        main_categories = ['proteins', 'starch', 'salad']
        
        if target_day == 'ALL_DAYS':
            # Use day-specific items for each day
            for day in DAYS_OF_WEEK:
                for category in main_categories:  # Only process main categories
                    if category in weekly_menu_items.get(day, {}):
                        # Replace existing items with new ones from PDF
                        weekly_menu[day][category] = weekly_menu_items[day][category][:]
                        items_added += len(weekly_menu_items[day][category])
        else:
            if target_day in DAYS_OF_WEEK:
                # Use items for the specific target day
                for category in main_categories:  # Only process main categories
                    if category in weekly_menu_items.get(target_day, {}):
                        # Replace existing items with new ones from PDF
                        weekly_menu[target_day][category] = weekly_menu_items[target_day][category][:]
                        items_added += len(weekly_menu_items[target_day][category])
        
        save_weekly_menu(weekly_menu)
        target_display = "the entire week" if target_day == 'ALL_DAYS' else target_day
        
        # Show breakdown of items added by day
        breakdown = []
        for day in DAYS_OF_WEEK:
            day_total = sum(len(items) for items in weekly_menu_items.get(day, {}).values())
            if day_total > 0:
                breakdown.append(f"{day}: {day_total} items")
        
        # Prepare success message
        success_msg = f'PDF processed successfully! Added menu items to {target_display}.'
        if breakdown:
            success_msg += f' {", ".join(breakdown)}'
            flash(success_msg)
        else:
            flash('PDF processed but no items were extracted for the selected day(s).')
        
    except ImportError:
        flash('PDF processing library not available. Please add menu items manually.')
    except Exception as e:
        print(f"PDF processing error: {str(e)}")
        flash(f'Error processing PDF: {str(e)}')
    
    return redirect(url_for('admin_page', section=current_section))

def process_pdf_text(text_content):
    """Process extracted PDF text and categorize menu items by day, filtering out irrelevant content"""
    lines = text_content.split('\n')
    
    # Initialize menu structure for each day - only main categories
    weekly_menu_items = {}
    for day in DAYS_OF_WEEK:
        weekly_menu_items[day] = {
            'proteins': [],
            'starch': [],
            'salad': []
        }
    
    # Keywords for categorization
    protein_keywords = ['chicken', 'beef', 'fish', 'pork', 'meat', 'turkey', 'lamb', 'seafood', 'prawns', 'shrimp', 'salmon', 'tuna', 'liver', 'oxtail', 'offals', 'stew', 'curry', 'mince', 'beans']
    starch_keywords = ['rice', 'pasta', 'bread', 'potato', 'noodles', 'quinoa', 'couscous', 'barley', 'oats', 'wheat', 'sadza', 'spaghetti', 'bolognaise', 'mashed', 'fried rice']
    salad_keywords = ['salad', 'greens', 'vegetables', 'lettuce', 'tomato', 'cucumber', 'spinach', 'kale', 'cabbage', 'coleslaw', 'chakalaka', 'nicoise', 'slaw', 'salsa', 'coslow']
    
    # Content to filter out
    filter_patterns = [
        r'takeaways.*catering', r'phone.*\+263', r'@mrsadza', r'compiled by',
        r'thank you', r'mr sadza', r'corporate lunch', r'harare',
        r'robson.*manyika', r'wayne.*street', r'pfugari.*house',
        r'weddings.*conferences', r'^\d+$', r'^\*+$', r'production$',
        r'bowls day', r'menu$'
    ]
    
    # Day patterns - more comprehensive
    day_patterns = {
        'monday': r'monday|mon\b',
        'tuesday': r'tuesday|tue\b|tues\b',
        'wednesday': r'wednesday|wed\b',
        'thursday': r'thursday|thu\b|thur\b|thurs\b',
        'friday': r'friday|fri\b'
    }
    
    # Category header patterns
    category_patterns = {
        'proteins': r'^protein[s]?$|^meat[s]?$|^main[s]?$',
        'starch': r'^starch[es]?$|^carb[s]?$|^side[s]?$',
        'salad': r'^salad[s]?$|^vegetable[s]?$|^veg[s]?$'
    }
    
    current_day = None
    current_category = None
    
    print(f"=== ENHANCED PDF PROCESSING DEBUG ===")
    
    for i, line in enumerate(lines):
        line = line.strip()
        if not line or len(line) < 3:
            continue
            
        line_lower = line.lower()
        
        # Check for day headers
        day_found = None
        for day_name, pattern in day_patterns.items():
            if re.search(pattern, line_lower):
                day_found = day_name.title()
                break
        
        if day_found and day_found in DAYS_OF_WEEK:
            current_day = day_found
            current_category = None
            print(f"Found day: {current_day} (line {i}: '{line}')")
            continue
        
        # Check for category headers
        category_found = None
        for category_name, pattern in category_patterns.items():
            if re.search(pattern, line_lower):
                category_found = category_name
                break
        
        if category_found:
            current_category = category_found
            print(f"Found category: {current_category} for day: {current_day} (line {i}: '{line}')")
            continue
        
        # Apply filter patterns
        should_skip = False
        for pattern in filter_patterns:
            if re.search(pattern, line_lower):
                should_skip = True
                break
        
        if should_skip:
            continue
        
        # Skip single words, numbers, or very short phrases
        if len(line.split()) < 2 or line.isdigit() or len(line) < 5:
            continue
        
        # Skip lines that are mostly numbers or special characters
        if re.match(r'^[\d\s\W]+$', line):
            continue
        
        # Clean up the line
        line = re.sub(r'\s+', ' ', line).strip()
        
        # If we have both day and category context, add the item
        if current_day and current_category:
            if line not in weekly_menu_items[current_day][current_category]:
                weekly_menu_items[current_day][current_category].append(line)
                print(f"Added to {current_day} {current_category}: {line}")
            continue
        
        # If we have a day but no category, try to auto-categorize
        if current_day:
            categorized = False
            
            # Check proteins first (most specific)
            for keyword in protein_keywords:
                if keyword in line_lower:
                    if line not in weekly_menu_items[current_day]['proteins']:
                        weekly_menu_items[current_day]['proteins'].append(line)
                        print(f"Auto-categorized to {current_day} proteins: {line}")
                        categorized = True
                        break
            
            if not categorized:
                # Check starch
                for keyword in starch_keywords:
                    if keyword in line_lower:
                        if line not in weekly_menu_items[current_day]['starch']:
                            weekly_menu_items[current_day]['starch'].append(line)
                            print(f"Auto-categorized to {current_day} starch: {line}")
                            categorized = True
                            break
            
            if not categorized:
                # Check salad
                for keyword in salad_keywords:
                    if keyword in line_lower:
                        if line not in weekly_menu_items[current_day]['salad']:
                            weekly_menu_items[current_day]['salad'].append(line)
                            print(f"Auto-categorized to {current_day} salad: {line}")
                            categorized = True
                            break
            
            # If not categorized but looks like food, add to proteins
            if not categorized:
                food_indicators = ['grilled', 'fried', 'baked', 'roasted', 'steamed', 'mixed', 'fresh', 'with', 'and', 'honey', 'garlic', 'herb', 'spicy', 'sweet']
                has_food_indicators = any(indicator in line_lower for indicator in food_indicators)
                
                if (len(line.split()) >= 2 and 
                    has_food_indicators and 
                    len(line) > 8 and
                    not line.lower().startswith(('the ', 'and ', 'or ', 'but ', 'a ', 'an '))):
                    if line not in weekly_menu_items[current_day]['proteins']:
                        weekly_menu_items[current_day]['proteins'].append(line)
                        print(f"Auto-categorized to {current_day} proteins (general): {line}")
    
    # Sort and clean up results
    for day in weekly_menu_items:
        for category in weekly_menu_items[day]:
            weekly_menu_items[day][category] = sorted(list(set(weekly_menu_items[day][category])))
    
    # Print summary
    total_items = sum(len(items) for day_menu in weekly_menu_items.values() for items in day_menu.values())
    print(f"Total items parsed: {total_items}")
    for day in DAYS_OF_WEEK:
        day_total = sum(len(items) for items in weekly_menu_items[day].values())
        if day_total > 0:
            print(f"{day}: {day_total} items")
            for category in ['proteins', 'starch', 'salad']:
                items = weekly_menu_items[day][category]
                if items:
                    print(f"  {category}: {len(items)} items - {items[:3]}{'...' if len(items) > 3 else ''}")
    
    print(f"=== END ENHANCED PDF PROCESSING DEBUG ===")
    
    return weekly_menu_items

if __name__ == '__main__':
    # Run initial cleanup on startup
    cleanup_old_orders()

    # Start the background cleanup scheduler
    start_cleanup_scheduler()

    # Get port from environment or default to 5000
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
